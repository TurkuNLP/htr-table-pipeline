from pathlib import Path
import re
from typing import Literal

import openpyxl

from postprocess.table_types import ParishBook, PrintType, TableAnnotation

# Functions for reading metadata from the annotations excel file


def extract_significant_parts_jpg(filename: str) -> dict[str, str] | None:
    """
    Extracts significant parts from a filename based on a specific pattern.

    The expected pattern is:
    autods-LOCATION_CATEGORY_YEARRANGE_DETAILS_NUMBER.xml

    Args:
        filename (str): The filename string.

    Returns:
        dict: A dictionary containing the extracted parts
              ("location", "category", "year_range", "details", "page_number")
              if the pattern matches, otherwise None.
    """
    # Regex breakdown:
    # ^mands-          : Starts with "mands-"
    # ([a-z_]+)        : Group 1 (location): lowercase letters and underscores
    # _                : Underscore separator
    # ([a-z_]+)        : Group 2 (category): lowercase letters and underscores
    # _                : Underscore separator
    # (\d{4}-\d{4})    : Group 3 (year_range): YYYY-YYYY
    # _                : Underscore separator
    # (.+)             : Group 4 (details): any characters (at least one)
    # _                : Underscore separator
    # (\d+)            : Group 5 (page_number): one or more digits
    # \.xml$           : Ends with ".xml"
    pattern = re.compile(r"^autods_([a-z_]+)_([a-z_]+)_(\d{4}-\d{4})_(.+)_(\d+)\.jpg$")
    match = pattern.match(filename)

    if match:
        parts = {
            "parish": match.group(1),
            "doctype": match.group(2),
            "year_range": match.group(3),
            "source": match.group(4),
            "page_number": match.group(5),
        }
        return parts
    else:
        return None


def extract_significant_parts_xml(filename: str) -> dict[str, str] | None:
    """
    Extracts significant parts from a filename based on a specific pattern.

    The expected pattern is:
    autods_PARISH_DOCTYPE_YEARRANGE_SOURCE_PAGENUMBER.xml
    or
    mands-PARISH_DOCTYPE_YEARRANGE_SOURCE_PAGENUMBER.xml

    Where:

    Args:
        filename (str): The filename string.

    Returns:
        dict: A dictionary containing the extracted parts
              ("parish", "doctype", "year_range", "source", "page_number")
              if the pattern matches, otherwise None.
    """
    # Regex breakdown:
    # ^mands-          : Starts with "autods_" or "mands-"
    # ([a-z_]+)        : Group 1 (parish): lowercase letters and underscores
    # _                : Underscore separator
    # ([a-z_]+)        : Group 2 (doctype): lowercase letters and underscores
    # _                : Underscore separator
    # (\d{4}-\d{4})    : Group 3 (year_range): YYYY-YYYY
    # _                : Underscore separator
    # (.+)             : Group 4 (source): any characters (at least one)
    # _                : Underscore separator
    # (\d+)            : Group 5 (page_number): one or more digits
    # \.xml$           : Ends with ".xml"
    pattern = re.compile(
        r"^(?:autods_|mands-)([a-z_]+)_([a-z_]+)_(\d{4}-\d{4})_(.+)_(\d+)\.xml$"
    )
    match = pattern.match(filename)

    if match:
        parts = {
            "parish": match.group(1),
            "doctype": match.group(2),
            "year_range": match.group(3),
            "source": match.group(4),
            "page_number": match.group(5),
        }
        return parts
    else:
        return None


def get_book_folder_id_for_xml(xml_file: Path) -> str:
    """
    Returns the folder ID for a given XML file based on its name.
    The expected pattern is:
    autods_LOCATION_CATEGORY_YEARRANGE_DETAILS_NUMBER.xml

    Args:
        xml_file (Path): The path to the XML file.

    Returns:
        str: The folder ID if the pattern matches, otherwise None.
    """
    xml_file_name = xml_file.name.lower()
    parts = extract_significant_parts_xml(xml_file_name)
    if parts is not None:
        return f"{parts['parish']}_{parts['doctype']}_{parts['year_range']}_{parts['source']}"
    raise ValueError(
        f"Could not extract significant parts from XML file name: {xml_file_name}"
    )


def get_print_type_str_for_jpg(jpg_file: Path, annotations_file: Path) -> str | None:
    """
    Returns the print type for a jpg file from the annotations file. Expensive utility function, do not use in production.
    """
    # Get the jpg file name without the extension
    jpg_file_name = jpg_file.name.lower()
    parts = extract_significant_parts_jpg(jpg_file_name)
    assert parts is not None, f"Could not extract parts from {jpg_file_name}"

    books = get_parish_books_from_annotations(annotations_file)

    for book in books:
        if (
            book.folder_id()
            == f"{parts['parish']}_{parts['doctype']}_{parts['year_range']}_{parts['source']}"
        ):
            print_type_str = book.get_type_for_opening(int(parts["page_number"]))
            return print_type_str
    return None


def get_print_type_str_for_xml(xml_file: Path, annotations_file: Path) -> str | None:
    """
    Returns the print type for a jpg file from the annotations file. Expensive utility function, do not use in production.
    """
    # Get the jpg file name without the extension
    jpg_file_name = xml_file.name.lower()
    parts = extract_significant_parts_xml(jpg_file_name)
    assert parts is not None, f"Could not extract parts from {jpg_file_name}"

    books = get_parish_books_from_annotations(annotations_file)

    for book in books:
        if (
            book.folder_id()
            == f"{parts['parish']}_{parts['doctype']}_{parts['year_range']}_{parts['source']}"
        ):
            print_type_str = book.get_type_for_opening(int(parts["page_number"]))
            return print_type_str
    return None


def get_print_type_for_jpg(jpg_file: Path, annotations_file: Path) -> PrintType:
    # TODO Currently used by table_corrector_agent.py, move to something that doesn't have to read the annotations file every call...
    print_type_str = get_print_type_str_for_jpg(jpg_file, annotations_file)
    print_dict = read_print_type_annotations(annotations_file)
    if print_type_str is None:
        raise ValueError(f"No print type found for {jpg_file}.")

    return print_dict[print_type_str.lower()]


def get_print_type_for_xml(xml_file: Path, annotations_file: Path) -> PrintType:
    # TODO Currently used by table_corrector_agent.py, move to something that doesn't have to read the annotations file every call...
    assert xml_file.suffix == ".xml", "The file must be an XML file."
    print_type_str = get_print_type_str_for_xml(xml_file, annotations_file)
    print_dict = read_print_type_annotations(annotations_file)
    if print_type_str is None:
        raise ValueError(f"No print type found for {xml_file}.")

    return print_dict[print_type_str.lower()]


def parse_book_types(raw_book_type: str) -> dict[str, tuple[int, int]]:
    """
    Parses the book type from the annotations file.

    E.g. Print 21:1-50, print 22:50-661 -> {"Print 21": (1, 50), "Print 22": (50, 661)}
    """
    book_type = {}
    for book in raw_book_type.split(","):
        book = book.strip()
        if ":" in book:
            name, years = book.split(":")
            years = years.split("-")
            start = int(years[0])
            end = int(years[1])
            book_type[name.strip().lower()] = (start, end)
        else:
            name = book
            book_type[name.strip().lower()] = (0, 999999)
    return book_type


def get_parish_books_from_annotations(annotations_file: Path) -> list[ParishBook]:
    """
    Returns a list of parish books from the annotations excel file.
    """
    wb = openpyxl.load_workbook(annotations_file)
    ws = wb.worksheets[0]  # first sheet

    parish_books = []
    for row in range(2, ws.max_row + 1):  # skip headers but read everything else
        parish_name = ws.cell(row=row, column=1).value
        if parish_name in ["", None]:
            break
        raw_book_type = str(ws.cell(row=row, column=19).value)
        years = ws.cell(row=row, column=8).value
        source = ws.cell(row=row, column=9).value
        doctype = ws.cell(row=row, column=17).value
        book = ParishBook(
            str(parish_name),
            parse_book_types(raw_book_type),
            str(years),
            str(source).lower(),
            str(doctype).lower(),
        )
        parish_books.append(book)

    return parish_books


def read_print_type_annotations(annotation_file: Path) -> dict[str, PrintType]:
    wb = openpyxl.load_workbook(annotation_file)
    ws = wb.worksheets[1]  # second sheet
    types: dict[str, PrintType] = {}
    for row in range(2, ws.max_row + 1):  # skip headers but read everything else
        print_type = str(ws.cell(row=row, column=1).value).lower()
        # table_type is either "one table" or "two tables" or None
        table_type: str | None = ws.cell(row=row, column=2).value  # type: ignore
        if table_type is None:
            continue  # Skip if table_type is second row of "two tables", the row is included in the first one
        direction = str(ws.cell(row=row, column=3).value).strip()
        if direction not in [
            "in",
            "out",
            "both",
            "out abroad",
        ]:  # both means separate columns for both
            direction = "?"
        number_of_columns_str = str(ws.cell(row=row, column=6).value)
        number_of_columns = int(
            float(
                number_of_columns_str.replace("left", "").replace("right", "").strip()
            )
        )
        # 1) one table per opening ("one table")
        if table_type == "one table":
            headers = []
            for i in range(7, 7 + number_of_columns):
                column_header = str(ws.cell(row=row, column=i).value)
                headers.append(column_header)

            type = PrintType(
                name=print_type,
                tables_per_jpg="one table",
                table_annotations=[
                    TableAnnotation(
                        print_type=print_type,
                        direction=direction,
                        col_headers=headers,
                        page="opening",
                    )
                ],
            )
            types[print_type] = type
            continue

        # 2) two tables per opening ("two tables")

        # left page
        left_headers: list[str] = []
        for i in range(7, 7 + number_of_columns):
            column_header = str(ws.cell(row=row, column=i).value)
            if isinstance(column_header, str):
                column_header = column_header.strip()
            left_headers.append(column_header)
        left_table = TableAnnotation(
            print_type=print_type,
            direction=direction,
            col_headers=left_headers,
            page="left",
        )

        # right page
        direction = str(ws.cell(row=row + 1, column=3).value).strip()
        if direction not in [
            "in",
            "out",
            "both",
            "out abroad",
        ]:  # both means separate columns for both
            direction = "?"
        number_of_columns_str = str(ws.cell(row=row + 1, column=6).value)
        number_of_columns = int(
            float(
                number_of_columns_str.replace("left", "").replace("right", "").strip()
            )
        )

        right_headers: list[str] = []
        for i in range(7, 7 + number_of_columns):
            column_header = str(ws.cell(row=row + 1, column=i).value)
            if isinstance(column_header, str):
                column_header = column_header.strip()
            right_headers.append(column_header)

        right_table = TableAnnotation(
            print_type=print_type,
            direction=direction,
            col_headers=right_headers,
            page="right",
        )

        # Add the print type
        type = PrintType(
            name=print_type,
            tables_per_jpg="two tables",
            table_annotations=[
                left_table,
                right_table,
            ],
        )
        types[print_type] = type
    return types


class BookAnnotationReader:
    def __init__(self, annotation_file: Path):
        """
        Initializes the BookAnnotationReader with the path to the annotations file.
        """
        self.annotation_file = annotation_file
        self.parish_books = get_parish_books_from_annotations(annotation_file)
        self.parish_book_mapping: dict[str, ParishBook] = {
            book.folder_id(): book for book in self.parish_books
        }
        self.print_type_mapping = read_print_type_annotations(annotation_file)

    def get_book(self, id: str) -> ParishBook:
        """
        Returns the ParishBook object for the given book ID.
        The ID is expected to be in the format 'parish_doctype_yearrange_source'.
        """
        if id not in self.parish_book_mapping:
            raise ValueError(f"Book with ID {id} not found in annotations.")
        return self.parish_book_mapping[id]

    def get_book_for_xml(self, xml_file: str) -> ParishBook:
        """
        Returns the ParishBook object for the given XML file.
        The XML file name is expected to follow the pattern:
        autods_parish_doctype_yearrange_source_page.xml
        """
        assert xml_file.endswith(".xml"), "The file must be an XML file."
        parts = extract_significant_parts_xml(xml_file.lower())
        if parts is None:
            raise ValueError(f"Could not extract significant parts from {xml_file}")
        return self.get_book(
            f"{parts['parish']}_{parts['doctype']}_{parts['year_range']}_{parts['source']}"
        )

    def get_print_type(self, print_type_str: str) -> PrintType:
        """
        Returns the PrintType object for the given print type string.
        The string is expected to be in lowercase.
        """
        if print_type_str not in self.print_type_mapping:
            raise ValueError(f"Print type {print_type_str} not found in annotations.")
        return self.print_type_mapping[print_type_str]

    def get_table_headers(
        self,
        book_id: str,  # e.g. "ahlainen_muuttaneet_1900-1910_ap_sis"
        opening: int,  # e.g. "1" or "2" for the first or second opening of the book
        page_side: Literal[
            "left", "right", "both"
        ] = "both",  # "both" means table covers both left and right pages
    ) -> list[str] | None:
        """
        Returns the column headers for the given opening and book.

        If the header's aren't known, returns None.
        """
        book = self.get_book(book_id)
        type_str = book.get_type_for_opening(opening)
        if "print" not in type_str.lower():
            return None
        print_type = self.get_print_type(type_str.lower())
        if page_side == "both":
            return print_type.table_annotations[0].col_headers
        if page_side == "left":
            return print_type.table_annotations[0].col_headers
        if page_side == "right":
            return print_type.table_annotations[-1].col_headers
        return None

    def get_table_direction(
        self,
        book_id: str,  # e.g. "ahlainen_muuttaneet_1900-1910_ap_sis"
        opening: int,  # e.g. "1" or "2" for the first or second opening of the book
        page_side: Literal[
            "left", "right", "both"
        ],  # "both" means table covers both left and right pages
    ) -> str:
        """
        Returns in/out for the given opening and book.
        """
        book = self.get_book(book_id)
        type_str = book.get_type_for_opening(opening)
        if "print" in type_str.lower():
            # Printed books
            print_type = self.get_print_type(type_str.lower())
            if page_side == "both":
                return print_type.table_annotations[0].direction
            if page_side == "left":
                return print_type.table_annotations[0].direction
            if page_side == "right":
                return print_type.table_annotations[-1].direction
        else:
            # Handwritten books
            # Use regex to find the direction in the book type
            return extract_in_out(type_str.lower())


def extract_in_out(text: str) -> str:
    """
    Extract in/out information from a string.
    Returns the in/out part if found, empty string otherwise.
    Preserves the order of in/out or out/in.
    """
    pattern = r"\b((?:in/out|out/in|in|out))$"
    match = re.search(pattern, text.lower().strip())
    return match.group(1) if match else ""
