# functions to load and process data
# annotated excel
# predicted zip files

import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import glob
import re
import csv

def read_layout_annotations(annotation_file):

    wb = openpyxl.load_workbook(annotation_file)
    ws = wb.worksheets[1] # second sheet
    print(wb)
    types = {} # dictionary, where print type is key and value is a ???
    for row in range(2, ws.max_row + 1): # skip headers but read everything else
        print_type = ws.cell(row=row, column=1).value
        table_type = ws.cell(row=row, column=2).value # either "one table" or "two tables"
        assert table_type in ["one table", "two tables", None], table_type
        direction = ws.cell(row=row, column=3).value 
        if direction not in ["in", "out", "both", "out abroad"]: # both means separate columns to both
            print("unknown direction", direction)
            direction = "unknown"
        number_of_columns = ws.cell(row=row, column=6).value
        if number_of_columns == "unknown":
            print("Skipping layout annotation for", print_type, " because it has unknown number of columns.")
            continue
        # 1) one table per opening ("one table")
        if table_type == "one table":
            number_of_columns = int(number_of_columns)
            headers = []
            for i in range(7, 7+number_of_columns):
                column_header = ws.cell(row=row, column=i).value
                headers.append(column_header)
            d = {"page": "opening", "direction": direction, "columns": number_of_columns, "headers": headers}
            types[print_type] = [d]
            continue

        # 2) two tables per opening ("two tables")
        # this is either left or right page layout
        page, number_of_columns = number_of_columns.split()
        number_of_columns = int(number_of_columns)

        headers = []
        for i in range(7, 7+number_of_columns):
            column_header = ws.cell(row=row, column=i).value
            if isinstance(column_header, str):
                column_header = column_header.strip()
            headers.append(column_header)

        if page == "left": #this is the first one, has all information present
            d = {"page": "left", "direction": direction, "columns": number_of_columns, "headers": headers}
            types[print_type] = [d]
        else: # this is the second one, take the layout from the first one
            assert page == "right"
            d = {"page": "right", "direction": direction, "columns": number_of_columns, "headers": headers}
            types[print_type].append(d)

    return types


def layout_unknown(layout_type):
    supported_types = ['free text', 'empty', 'wrong', 'digital', 'handrawn', 'halfbook', "print 13a", "print 13b"]
    for i in range(1, 56): # add all known prints
        supported_types.append(f"print {i}")
    for l in ["in", "out", "in/out", "out/in", "misc", "abroad"]: # add all known variants
        supported_types.append(f"handrawn {l}")
        supported_types.append(f"halfbook {l}")
        supported_types.append(f"free text {l}")
    if layout_type in supported_types:
        return False
    return True

def is_same_as(example):
    # Same as charecters(", D, Do)
    same_as_patterns = [r'^"+$', r'^D\.?$', r'^Do\.?$', r'^d\.?$', r'^do\.?$', r'^Sn\.?$', r'^Sm\.?$', r'^sn\.?$', r'^sm\.?$']
    example = re.sub(r'\s+', '', example)
    #print(anno_pred_pairs)
    same_as = any(re.match(pattern, example) for pattern in same_as_patterns)
    return same_as

def repeat_column(table, row_idx, column_idx):
    if is_same_as(table[row_idx][column_idx]):
        repeated = None
        for i in range(row_idx-1, -1, -1):
            if not is_same_as(table[i][column_idx]) and table[i][column_idx] != "---" and table[i][column_idx] != "":
                repeated = table[i][column_idx]
                break
        if repeated:
            return repeated
        else:
            return f"Not able to resolve: {table[row_idx][column_idx]}"
    return table[row_idx][column_idx]


def repeat_table_data(table):
    for row_idx in range(0, len(table)):
        for column_idx in range(0, len(table[row_idx])):
            table[row_idx][column_idx] = repeat_column(table, row_idx, column_idx)
    return table

def read_text_from_cell(cell): # TODO: how to combine multiple text lines?
    # cell may have multiple text lines
    text_lines = []
    for line in cell.findall('.//ns:TextLine', namespace):
        text_line = line.find('.//ns:TextEquiv', namespace).find('.//ns:Unicode', namespace).text
        if text_line:
            text_lines.append(text_line)
    text = "\n".join(l for l in text_lines).strip()
    if text == "":
        text = "---"
    return text

def yield_parish_files_and_annotations(annotation_file, warn=False):
    # read book annotations from excel and yield books, where book is a list of pages
    wb = openpyxl.load_workbook(annotation_file)
    ws = wb.worksheets[0] # first sheet
    #ws = wb.active
    
    for row in range(2, ws.max_row + 1): # skip headers but read everything else, each row is one book
        if all([cell.value == None for cell in ws[row]]):
            print(f"Reading parish annotations done, read {row-1} rows.")
            break
        normalized_parish = ws.cell(row=row, column=1).value
        number_of_images = int(ws.cell(row=row, column=4).value)
        book_type = ws.cell(row=row, column=5).value # always muuttaneet?
        book_url = ws.cell(row=row, column=6).value
        years = ws.cell(row=row, column=8).value
        scan_type = str(ws.cell(row=row, column=9).value).lower()
        annotations = ws.cell(row=row, column=19).value.lower()
        annotations = annotations.split(",")
        annotations = [n.strip() for n in annotations]

        # page number, file path, url, layout type
        
        pages = [{"page number": i,
                "file path": f"images/{normalized_parish}/muuttaneet_{years}_{scan_type}/pageText/autods_{normalized_parish}_muuttaneet_{years}_{scan_type}_{i}.xml",
                "url": f"https://www.sukuhistoria.fi/sshy/kirjat/Kirkonkirjat/{normalized_parish}/muuttaneet_{years}_{scan_type}/{i}.htm",
                "layout type": "unknown"}
                 for i in range(1, number_of_images+1)] # initialize with unknown, then fill in annotations (pages with missing annotations will remain uknown)
        
        book_name = f"autods_{normalized_parish}_muuttaneet_{years}_{scan_type}"

        # add layout types from annotations
        for annotation in annotations:
            if ":" in annotation: # includes page range
                layout_type, page_annotation = annotation.split(":")
                page_start, page_end = int(page_annotation.split('-')[0]), int(page_annotation.split('-')[1])+1
            else: # full book same layout
                layout_type = annotation
                page_start, page_end = 1, number_of_images+1

            # assert that we do have have the same page already annotated
            assert all([pages[i]["layout type"] == "unknown" for i in range(page_start-1, page_end-1)]), annotations

            layout_type = layout_type.strip()
            assert layout_unknown(layout_type) == False, layout_type
            for page_number in range(page_start, page_end):
                pages[page_number-1]["layout type"] = layout_type
                
        
        unknown_pages = sum([1 for p in pages if p["layout type"] == "unknown"])
        if warn and unknown_pages > 0:
            print("Warning", row, "has", unknown_pages, "unknown pages.", annotations)
        
        yield book_name, pages


namespace = {'ns': 'http://schema.primaresearch.org/PAGE/gts/pagecontent/2013-07-15'}

def extract_datatables_from_xml(xml_file):
    # read the xml file (predictions for one image), and return datatables
    tree = ET.parse(xml_file)
    root = tree.getroot()
    tables = []
    for table in root.findall('.//ns:TableRegion', namespace): # iterate over tables
        if len(table.findall('.//ns:TableCell', namespace)) < 2: # skip tables with only one cell, these are headers, or totally empty tables (no rows or columns)
            continue
        table_rows = []
        current_row = []
        current_row_id = None
        for cell in table.findall('.//ns:TableCell', namespace):
            row_id = int(cell.attrib.get("row"))
            if current_row_id != None and row_id != current_row_id: # previous row ended
                table_rows.append(current_row)
                current_row = []
                assert row_id == current_row_id + 1 # assert that these are in correct order 
            current_row_id = row_id
            column_id = int(cell.attrib.get("col"))
            assert int(column_id) == len(current_row) # assert that these are in correct order
            text = read_text_from_cell(cell)
            current_row.append(text)
        if current_row:
            table_rows.append(current_row)

        # repeat and otherwise postprocess the data
        table_rows = repeat_table_data(table_rows)
        tables.append(table_rows)

    return tables