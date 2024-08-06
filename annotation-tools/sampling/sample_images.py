from random import sample
import openpyxl
import re
import argparse
import os
import shutil
import zipfile

file_dir = "/scratch/project_2005072/moving_records_htr"  # muokkaa tämä  "/scratch/project_2005072/moving_records_htr"

def load_images_to_exclude(file_path):
    images_to_exclude = set()
    with open(file_path, 'r') as file:
        for line in file:
            parts = line.split()
            if parts:  # Check if the line is not empty
                image_path = parts[0]  
                images_to_exclude.add(image_path)
    return images_to_exclude

def yield_images(file_catalogy, layout_to_sample, images_to_exclude):
    wb = openpyxl.load_workbook(args.file_catalogy)

    ws = wb.active

    for row in range(2, ws.max_row + 1): 
        a_value = ws.cell(row=row, column=1).value
        d_value = ws.cell(row=row, column=4).value
        e_value = ws.cell(row=row, column=5).value
        f_value = ws.cell(row=row, column=6).value
        h_value = ws.cell(row=row, column=8).value
        i_value = ws.cell(row=row, column=9).value
        r_value = ws.cell(row=row, column=18).value

        parish_nor = a_value
        images = d_value
        doc = e_value
        book_url = f_value
        years = h_value
        source = i_value.lower()
        notes = r_value.lower()

        notes = notes.split(",")
        notes = [n.strip() for n in notes]
        for note in notes:
            if ":" in note:
                print_type, pages = note.split(":")     
            else:
                print_type = note
                pages = None

            all_printed = print_type.startswith('print')
            all_handdrawn = print_type.startswith(('handrawn', 'halfbook', 'free text'))  
            if (args.layout_to_sample == 'all printed' and all_printed) or print_type == args.layout_to_sample or (args.layout_to_sample == 'all handdrawn' and all_handdrawn):
                if pages:
                    page_ranges = pages.split(',')
                    for page_range in page_ranges:
                        if '-' in page_range:
                            start, end = map(int, page_range.split('-'))
                            for i in range(start, end):
                                file_path = f"images/{parish_nor}/{doc}_{years}_{source}/{parish_nor}_{doc}_{years}_{source}_{str(i+1)}.jpg" # <---  MUOKKAA TÄMÄ PUHDILLE   f"images/{parish_nor}/{doc}_{years}_{source}/{parish_nor}_{doc}_{years}_{source}_{str(i+1)}.jpg"
                                if file_path not in images_to_exclude:
                                    yield file_path, note
                        else:
                            continue
                else:
                    for i in range(images):
                        file_path = f"images/{parish_nor}/{doc}_{years}_{source}/{parish_nor}_{doc}_{years}_{source}_{str(i+1)}.jpg" # <---- MUOKKAA TÄMÄ PUHDILLE      f"images/{parish_nor}/{doc}_{years}_{source}/{parish_nor}_{doc}_{years}_{source}_{str(i+1)}.jpg"
                        if file_path not in images_to_exclude:
                            yield file_path, note
            else:
                continue
    


def main(args):
    if args.exclude_file:
        images_to_exclude = load_images_to_exclude(args.exclude_file)
    else:
        images_to_exclude = set()

    images_to_sample = list(yield_images(args.file_catalogy, args.layout_to_sample, images_to_exclude))
    sampled_images = sample(images_to_sample, args.sample_size)


    output_dir = os.path.join(file_dir, args.output_dir)
    if not os.path.exists(output_dir):
        print("Creating dir:", output_dir)
        os.makedirs(output_dir)

    readme_path = os.path.join(output_dir, 'README.txt')
    with open(readme_path, 'w') as f:
        for file, note in sampled_images:
            print_type = note.split(':')[0].strip() 
            f.write(f'{file}  {print_type}\n')

    with zipfile.ZipFile(args.zip_file, 'r') as zip_ref:
        for file, note in sampled_images:
            source_path = file
            destination_path = os.path.join(output_dir, os.path.basename(file))
            with zip_ref.open(source_path) as source, open(destination_path, 'wb') as dest:
                shutil.copyfileobj(source, dest)
  

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Majority vote')
    parser.add_argument('--file-catalogy', type=str, help='Big excel file with all the books')
    parser.add_argument('--layout-to-sample', type=str, required=True, help='Name of the annotated layout to sample (e.g. "print 1")')
    parser.add_argument('--sample-size', type=int, required=True, help='Number of samples to be taken')
    parser.add_argument('--output-dir', type=str, required=True, help='Directory to copy the sampled files')
    parser.add_argument('--exclude-file', type=str, required=False, help='Text file with list of images to exclude')
    parser.add_argument('--zip-file', type=str, required=False, help='Zip file containing the images')
    args = parser.parse_args()
    
    main(args)

    # run: python sample_images.py --file-catalogy excel.xlsx --layout-to-sample "print 1" --sample-size 100 --output-dir sample-print1 --exclude-file sample1.txt --zip-file images.zip