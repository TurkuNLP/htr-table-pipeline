from random import sample
import openpyxl
import re
import argparse
import os
import shutil
import zipfile
from collections import Counter
import xml.etree.ElementTree as ET

namespace = {'ns': 'http://schema.primaresearch.org/PAGE/gts/pagecontent/2013-07-15'}

def yield_images(file_catalogy):
    wb = openpyxl.load_workbook(file_catalogy)
    ws = wb.active
    for row in range(2, ws.max_row + 1): 
        parish_nor = ws.cell(row=row, column=1).value # normalized parish name
        images = ws.cell(row=row, column=4).value # number of images in this book
        doc = ws.cell(row=row, column=5).value # muuttaneet
        book_url = ws.cell(row=row, column=6).value # url
        years = ws.cell(row=row, column=8).value # book years
        source = ws.cell(row=row, column=9).value.lower() # AP/MKO etc.
        annotations = ws.cell(row=row, column=18).value.lower() # annotation

        notes = annotations.split(",")
        notes = [n.strip() for n in notes]
        annotated_types = {page_number: None for page_number in range(1, images+1)}
        #print(annotated_types)
        for i, note in enumerate(notes):
            if ":" in note:
                print_type, pages = note.split(":")
                if "-" not in pages:  #something weird in annotation
                    pages = None
            else:
                print_type = note
                pages = None

            if print_type.startswith(("print ", "handrawn", "halfbook", "free text")):  # this is some kind of file (print, handdrawn, halfbook, free text or other)
                #print(print_type, pages, annotations)
                if pages is not None:
                    start, end = pages.split('-')
                    for i in range(int(start), int(end)+1):
                        annotated_types[i] = print_type
                else:  # all pages follow thid print type
                    for page_number in range(1, images+1):
                        annotated_types[page_number] = print_type   
        
        # all notes processed, yield pages with layout type
        #print(annotated_types)
        for page_number, layout_type in annotated_types.items():
            file_path = f"{parish_nor}_{doc}_{years}_{source}_{str(page_number)}.xml"
            if layout_type is None:
                layout_type = "other layout"
            #print(file_path, layout_type)
            yield file_path, layout_type

# Function, that reads local files that have been cloned from htr-annotations GitHub repository.        
def read_local_files(directory):
    local_files = []
    cell_annotated_local_files = []
    total_text_lines = 0
    total_unclear_lines = 0


    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(".xml"):
                if file.lower().startswith("mands-"):
                    excel_name = os.path.basename(file).removeprefix("mands-")
                    local_files.append(excel_name)
                else:
                    local_files.append(os.path.basename(file))

                file_path = os.path.join(root, file)

            
                tree = ET.parse(file_path)
                root_element = tree.getroot()
            
                # this is for finding the textlines in the xml-files
                for element in root_element.iter():
                    text_line = element.attrib.get('custom', '')
                    if "readingOrder" in text_line:
                        unicode_elem = element.find('.//ns:Unicode', namespace)
                        if unicode_elem is not None and unicode_elem.text:
                            if "?" in {unicode_elem.text.strip()}:
                                total_unclear_lines += 1
                            else:
                                total_text_lines += 1
                        else:
                            continue

                # checks if the xml-files have been cell-annotated
                for element in root_element.iter():
                    custom_attrib = element.attrib.get('custom', '')
                    if "structure" in custom_attrib:
                        if file.lower().startswith("mands-"):
                            excel_name = os.path.basename(file).removeprefix("mands-")
                            cell_annotated_local_files.append(excel_name)
                        else:
                            cell_annotated_local_files.append(os.path.basename(file))
                        break

    return local_files, cell_annotated_local_files, total_text_lines, total_unclear_lines

def main(args):
    images = list(yield_images(args.file_catalogy))  
    total_images_in_excel = len(images)
    print(f"Total images in Excel: {total_images_in_excel}") 

    excel_layout_counter = Counter([layout for file_path, layout in images])

    total_printed_in_excel = sum(count for layout, count in excel_layout_counter.items() if layout.startswith("print"))
    total_handdrawn_in_excel = sum(count for layout, count in excel_layout_counter.items() if layout.startswith("handrawn"))
    total_halfbook_in_excel = sum(count for layout, count in excel_layout_counter.items() if layout.startswith("halfbook"))
    total_free_text_in_excel = sum(count for layout, count in excel_layout_counter.items() if layout.startswith("free text"))
    total_other_layout_in_excel = total_images_in_excel - (total_printed_in_excel + total_handdrawn_in_excel + total_halfbook_in_excel + total_free_text_in_excel)

    print(f"Total printed in Excel: {total_printed_in_excel} ({round((total_printed_in_excel / total_images_in_excel) * 100, 2)}%)")
    print(f"Total handdrawn in Excel: {total_handdrawn_in_excel} ({round((total_handdrawn_in_excel / total_images_in_excel) * 100, 2)}%)")
    print(f"Total halfbook in Excel: {total_halfbook_in_excel} ({round((total_halfbook_in_excel / total_images_in_excel) * 100, 2)}%)")
    print(f"Total free text in Excel: {total_free_text_in_excel} ({round((total_free_text_in_excel / total_images_in_excel) * 100, 2)}%)")
    print(f"Total other layout in Excel: {total_other_layout_in_excel} ({round((total_other_layout_in_excel / total_images_in_excel) * 100, 2)}%)")

    local_files, cell_annotated_local_files, total_text_lines, total_unclear_lines = read_local_files(args.local_directory) 
    annotated_local_files = [file_path for file_path, layout in images if file_path in local_files] # muoattu tästä: [file_path for file_path, layout in images if file_path in local_files]
    cell_annotated_local_files = [file_path for file_path, layout in images if file_path in cell_annotated_local_files] # muokattu tästä: [file_path for file_path, layout in images if file_path in cell_annotated_local_files]
    total_annotated_files = len(annotated_local_files)
    total_cell_annotated_files = len(cell_annotated_local_files)

    local_layout_counter = Counter([layout for file_path, layout in images if file_path in local_files])
    cell_annotated_local_file_counter = Counter([layout for file_path, layout in images if file_path in cell_annotated_local_files])

    total_annotated_printed = sum(count for layout, count in local_layout_counter.items() if layout.startswith("print"))
    total_annotated_handdrawn = sum(count for layout, count in local_layout_counter.items() if layout.startswith("handrawn"))
    total_annotated_halfbook = sum(count for layout, count in local_layout_counter.items() if layout.startswith("halfbook"))
    total_annotated_free_text = sum(count for layout, count in local_layout_counter.items() if layout.startswith("free text"))
    total_annotated_other = total_annotated_files - (total_annotated_printed + total_annotated_handdrawn + total_annotated_halfbook + total_annotated_free_text)
    total_cell_annotated_files = sum(count for layout, count in cell_annotated_local_file_counter.items())
    total_cell_annotated_printed = sum(count for layout, count in cell_annotated_local_file_counter.items() if layout.startswith("print"))
    total_cell_annotated_handdrawn = sum(count for layout, count in cell_annotated_local_file_counter.items() if layout.startswith("handrawn"))
    total_cell_annotated_halfbook = sum(count for layout, count in cell_annotated_local_file_counter.items() if layout.startswith("halfbook"))
    total_cell_annotated_free_text = sum(count for layout, count in cell_annotated_local_file_counter.items() if layout.startswith("free text"))
    total_cell_annotated_other = total_cell_annotated_files - (total_cell_annotated_printed + total_cell_annotated_handdrawn + total_cell_annotated_halfbook + total_cell_annotated_free_text)

    print(f"Total annotated files: {total_annotated_files} ({round((total_annotated_files / total_images_in_excel) * 100, 2)}%)")
    print(f"Total annotated printed: {total_annotated_printed} ({round((total_annotated_printed / total_annotated_files) * 100, 2)}%)")
    print(f"Total annotated handdrawn: {total_annotated_handdrawn} ({round((total_annotated_handdrawn / total_annotated_files) * 100, 2)}%)")
    print(f"Total annotated halfbook: {total_annotated_halfbook} ({round((total_annotated_halfbook / total_annotated_files) * 100, 2)}%)")
    print(f"Total annotated free text: {total_annotated_free_text} ({round((total_annotated_free_text / total_annotated_files) * 100, 2)}%)")
    print(f"Total annotated other: {total_annotated_other} ({round((total_annotated_other / total_annotated_files) * 100, 2)}%)")
    print(f"Total cell annotated: {total_cell_annotated_files}({round((total_cell_annotated_files / total_annotated_files) * 100, 2)}%)")
    print(f"Total cell annotated printed: {total_cell_annotated_printed} ({round((total_cell_annotated_printed / total_cell_annotated_files) * 100, 2)}%)")
    print(f"Total cell annotated handdrawn: {total_cell_annotated_handdrawn} ({round((total_cell_annotated_handdrawn / total_cell_annotated_files) * 100, 2)}%)")
    print(f"Total cell annotated halfbook: {total_cell_annotated_halfbook} ({round((total_cell_annotated_halfbook / total_cell_annotated_files) * 100, 2)}%)")
    print(f"Total cell annotated free text: {total_cell_annotated_free_text} ({round((total_cell_annotated_free_text / total_cell_annotated_files) * 100, 2)}%)")
    print(f"Total cell annotated other: {total_cell_annotated_other} ({round((total_cell_annotated_other / total_cell_annotated_files) * 100, 2)}%)")

    print(f"Total annotated text lines: {total_text_lines}")
    print(f"Unclear annotated text lines (with '?'): {total_unclear_lines}")
    print(f"Percentage unclear: {round((total_unclear_lines / total_text_lines) * 100, 2)}%" if total_text_lines > 0 else "N/A")

    if args.verbose:  
        print("\n### Layout Statistics for Entire Excel ###")
        for layout, count in excel_layout_counter.most_common(100):
            print(f"LAYOUT: {layout}, PAGE COUNT: {count} ({round((count/len(images))*100, 2)}%)")
        
        print(f"\nTotal local files: {len(local_files)}")
        print("\n### Layout Statistics for Annotated Files ###")
        for layout, count in local_layout_counter.most_common(100):
            print(f"LAYOUT: {layout}, PAGE COUNT: {count} ({round((count/len(annotated_local_files))*100, 2)}%)")

        print("\n### Layout Statistics for Cell Annotated Files ###")
        for layout, count in cell_annotated_local_file_counter.most_common(100):
            print(f"LAYOUT: {layout}, PAGE COUNT: {count} ({round((count/len(cell_annotated_local_files))*100, 2)}%)")    


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Statistics for annotated images.')
    parser.add_argument('--file-catalogy', type=str, help='Excel file with book annotations.')
    parser.add_argument('--local-directory', type=str, help='Local directory where cloned files are stored.')
    parser.add_argument('--verbose', action='store_true', help='Increase output information.') 
    args = parser.parse_args()
    
    main(args)

# run: python stats.py --file-catalogy excel.xlsx --local-directory (path to cloned repository) e.g /home/work/htr-annotations --verbose (for more information)
